from openpyxl import load_workbook


__all__ = ['read_excel']


def read_excel(file):
    wb = load_workbook(filename=file)
    ws = wb.active
    url_list = []
    for row in range(1, ws.max_row + 1):
        for subfix in ['', 'www.']:
            url = ws[f'b{row}'].value

            if isinstance(url, str):
                if not url.startswith('http'):
                    url = f'{subfix}{url}'
                    url = f'http://{url}'
                if not url.endswith('/'):
                    url = f'{url}/'
                url_list.append(url)
    return list(set(url_list))
